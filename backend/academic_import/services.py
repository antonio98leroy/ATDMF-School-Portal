from decimal import Decimal, InvalidOperation
import re

from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from academics.models import Enrollment, Subject
from examinations.models import ResultPeriod, SubjectResult
from students.models import Student


PERIOD_CODES = [
    ResultPeriod.Code.FIRST_PERIOD,
    ResultPeriod.Code.SECOND_PERIOD,
    ResultPeriod.Code.THIRD_PERIOD,
    ResultPeriod.Code.FIRST_SEMESTER_EXAM,
    ResultPeriod.Code.FOURTH_PERIOD,
    ResultPeriod.Code.FIFTH_PERIOD,
    ResultPeriod.Code.SIXTH_PERIOD,
    ResultPeriod.Code.SECOND_SEMESTER_EXAM,
]


def clean_text(value):
    if value is None:
        return ""
    return " ".join(str(value).replace("\\n", " ").split()).strip()


def subject_code(name):
    words = re.findall(r"[A-Za-z0-9]+", name.upper())
    base = "".join(word[:3] for word in words)[:18] or "SUBJECT"
    code = base
    counter = 1
    while Subject.objects.filter(code=code).exclude(name__iexact=name).exists():
        counter += 1
        code = f"{base[:14]}{counter:03d}"
    return code


def as_score(value):
    if value in (None, ""):
        return None
    if isinstance(value, str) and value.startswith("="):
        return None
    try:
        score = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError):
        return None
    return score.quantize(Decimal("0.01"))


def match_enrollment(student_id, academic_year, class_section):
    student_id = clean_text(student_id)
    if not student_id:
        return None

    student = Student.objects.filter(
        admission_number__iexact=student_id
    ).first()

    if not student:
        return None

    return Enrollment.objects.filter(
        student=student,
        academic_year=academic_year,
        class_section=class_section,
    ).first()


def parse_academic_workbook(file_obj, academic_year, class_section):
    workbook = load_workbook(file_obj, data_only=True)

    required = {
        "1ST SEM-PRD-ASS",
        "2ND SEM-PRD-ASS",
        "MARKSHEET",
    }
    missing = sorted(required.difference(workbook.sheetnames))
    if missing:
        raise ValueError(
            "Missing required worksheet(s): " + ", ".join(missing)
        )

    first_sheet = workbook["1ST SEM-PRD-ASS"]
    second_sheet = workbook["2ND SEM-PRD-ASS"]
    marksheet = workbook["MARKSHEET"]

    periods = {}
    for code, _name in ResultPeriod.Code.choices:
        period, _ = ResultPeriod.objects.get_or_create(
            academic_year=academic_year,
            code=code,
            defaults={
                "name": dict(ResultPeriod.Code.choices)[code],
                "order": ResultPeriod.PERIOD_ORDER[code],
                "score_entry_open": True,
            },
        )
        periods[code] = period

    subject_starts = []
    for column in range(3, first_sheet.max_column + 1, 18):
        name = clean_text(first_sheet.cell(6, column).value)
        if not name or name == "0":
            continue
        subject_starts.append((column, name))

    rows = []
    for excel_row in range(9, max(first_sheet.max_row, second_sheet.max_row, marksheet.max_row) + 1):
        student_id = clean_text(marksheet.cell(excel_row, 2).value)
        if not student_id:
            continue

        enrollment = match_enrollment(
            student_id,
            academic_year,
            class_section,
        )

        for block_index, (assessment_col, subject_name) in enumerate(subject_starts):
            mark_col = 3 + (block_index * 8)

            subject, _ = Subject.objects.get_or_create(
                name__iexact=subject_name,
                defaults={
                    "name": subject_name,
                    "code": subject_code(subject_name),
                    "description": "Created from historical academic import.",
                },
            )

            period_sources = [
                (first_sheet, assessment_col, PERIOD_CODES[0]),
                (first_sheet, assessment_col + 6, PERIOD_CODES[1]),
                (first_sheet, assessment_col + 12, PERIOD_CODES[2]),
                (marksheet, mark_col + 3, PERIOD_CODES[3]),
                (second_sheet, assessment_col, PERIOD_CODES[4]),
                (second_sheet, assessment_col + 6, PERIOD_CODES[5]),
                (second_sheet, assessment_col + 12, PERIOD_CODES[6]),
                (marksheet, mark_col + 7, PERIOD_CODES[7]),
            ]

            for source_sheet, start_col, period_code in period_sources:
                errors = []
                values = {}

                if period_code in {
                    ResultPeriod.Code.FIRST_SEMESTER_EXAM,
                    ResultPeriod.Code.SECOND_SEMESTER_EXAM,
                }:
                    exam_score = as_score(
                        source_sheet.cell(excel_row, start_col).value
                    )
                    if exam_score is None:
                        continue
                    if exam_score < 0 or exam_score > 100:
                        errors.append("Semester examination score must be between 0 and 100.")
                    values = {
                        "assignment_score": Decimal("0"),
                        "class_activity_score": Decimal("0"),
                        "quiz_score": Decimal("0"),
                        "period_test_score": Decimal("0"),
                        "semester_exam_score": exam_score,
                    }
                else:
                    class_activity = as_score(source_sheet.cell(excel_row, start_col).value)
                    assignment = as_score(source_sheet.cell(excel_row, start_col + 1).value)
                    quiz = as_score(source_sheet.cell(excel_row, start_col + 2).value)
                    test = as_score(source_sheet.cell(excel_row, start_col + 3).value)

                    if all(value is None for value in [class_activity, assignment, quiz, test]):
                        continue

                    class_activity = class_activity or Decimal("0")
                    assignment = assignment or Decimal("0")
                    quiz = quiz or Decimal("0")
                    test = test or Decimal("0")

                    if not Decimal("0") <= class_activity <= Decimal("10"):
                        errors.append("Class participation must be between 0 and 10.")
                    if not Decimal("0") <= assignment <= Decimal("10"):
                        errors.append("Assignment must be between 0 and 10.")
                    if not Decimal("0") <= quiz <= Decimal("30"):
                        errors.append("Quiz must be between 0 and 30.")
                    if not Decimal("0") <= test <= Decimal("50"):
                        errors.append("Period test must be between 0 and 50.")

                    values = {
                        "assignment_score": assignment,
                        "class_activity_score": class_activity,
                        "quiz_score": quiz,
                        "period_test_score": test,
                        "semester_exam_score": Decimal("0"),
                    }

                if not enrollment:
                    errors.append(
                        f"No enrollment found for student ID {student_id} "
                        f"in {class_section} and {academic_year}."
                    )

                duplicate = False
                if enrollment:
                    duplicate = SubjectResult.objects.filter(
                        enrollment=enrollment,
                        subject=subject,
                        period=periods[period_code],
                    ).exists()

                rows.append({
                    "excel_row": excel_row,
                    "student_id": student_id,
                    "student_name": (
                        enrollment.student.full_name if enrollment else ""
                    ),
                    "enrollment_id": enrollment.id if enrollment else None,
                    "subject_id": subject.id,
                    "subject_name": subject.name,
                    "period_id": periods[period_code].id,
                    "period_code": period_code,
                    "period_name": periods[period_code].name,
                    "assignment_score": str(values["assignment_score"]),
                    "class_activity_score": str(values["class_activity_score"]),
                    "quiz_score": str(values["quiz_score"]),
                    "period_test_score": str(values["period_test_score"]),
                    "semester_exam_score": str(values["semester_exam_score"]),
                    "duplicate": duplicate,
                    "errors": errors,
                    "valid": not errors,
                })

    return rows


@transaction.atomic
def import_preview_rows(batch, user):
    created_ids = []
    imported_count = 0

    for row in batch.preview_data:
        if not row.get("valid") or row.get("duplicate"):
            continue

        result, created = SubjectResult.objects.get_or_create(
            enrollment_id=row["enrollment_id"],
            subject_id=row["subject_id"],
            period_id=row["period_id"],
            defaults={
                "assignment_score": Decimal(row["assignment_score"]),
                "class_activity_score": Decimal(row["class_activity_score"]),
                "quiz_score": Decimal(row["quiz_score"]),
                "period_test_score": Decimal(row["period_test_score"]),
                "semester_exam_score": Decimal(row["semester_exam_score"]),
                "remarks": (
                    f"Historical import: {batch.original_filename}, "
                    f"Excel row {row['excel_row']}."
                ),
                "entered_by": user,
                "approved": False,
                "published": False,
            },
        )

        if created:
            created_ids.append(result.id)
            imported_count += 1

    batch.created_result_ids = created_ids
    batch.imported_rows = imported_count
    batch.status = batch.Status.IMPORTED
    batch.imported_at = timezone.now()
    batch.save(
        update_fields=[
            "created_result_ids",
            "imported_rows",
            "status",
            "imported_at",
        ]
    )
    return imported_count


@transaction.atomic
def rollback_import(batch):
    deleted, _ = SubjectResult.objects.filter(
        id__in=batch.created_result_ids
    ).delete()

    batch.status = batch.Status.ROLLED_BACK
    batch.rolled_back_at = timezone.now()
    batch.save(
        update_fields=[
            "status",
            "rolled_back_at",
        ]
    )
    return deleted
