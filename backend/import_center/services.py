import re
from datetime import date, datetime, timedelta
from decimal import Decimal

from django.apps import apps
from django.db import transaction
from django.utils import timezone
from openpyxl import load_workbook

from academics.models import ClassSection, Enrollment, GradeLevel
from finance.models import Sponsor, StudentSponsorship
from students.models import Guardian, Student

from .models import ImportedObject


HEADER_ALIASES = {
    "student_name": {"student name", "name of student", "student"},
    "status": {"status", "new/old", "student status"},
    "gender": {"gender", "sex"},
    "date_of_birth": {"student (date of birth)", "date of birth", "dob", "birth date"},
    "guardian_name": {"parent/guardian name", "parent guardian name", "guardian", "parent name"},
    "guardian_address": {"parent address", "guardian address", "address"},
    "sponsor_name": {"us sponsors", "us sponsor", "sponsor", "sponsor name"},
    "grade": {"grade", "class", "admission class"},
    "comments": {"comments", "remarks", "notes"},
}


def clean_text(value):
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").split()).strip()


def normalize_header(value):
    return clean_text(value).lower().rstrip(":")


def parse_excel_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        return date(1899, 12, 30) + timedelta(days=int(value))
    text = clean_text(value)
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def normalize_gender(value):
    value = clean_text(value).lower()
    if value in {"m", "male", "boy"}:
        return "M"
    if value in {"f", "female", "girl"}:
        return "F"
    return ""


def normalize_status(value):
    value = clean_text(value).lower()
    if value in {"new", "n", "new student"}:
        return "NEW"
    return "RETURNING"


def normalize_grade(value):
    text = clean_text(value).upper().replace("KINDERGARTEN", "K").replace("KINDERGARDEN", "K")
    text = re.sub(r"\s+", " ", text)
    compact = re.sub(r"[^A-Z0-9]", "", text)
    if compact in {"K1", "KG1", "KINDER1"}:
        return "K-1", 1
    if compact in {"K2", "KG2", "KINDER2"}:
        return "K-2", 2
    word_numbers = {
        "ONE": 1, "TWO": 2, "THREE": 3, "FOUR": 4, "FIVE": 5, "SIX": 6,
        "SEVEN": 7, "EIGHT": 8, "NINE": 9, "TEN": 10, "ELEVEN": 11, "TWELVE": 12,
    }
    match = re.search(r"(\d{1,2})", text)
    number = int(match.group(1)) if match else None
    if number is None:
        for word, value in word_numbers.items():
            if word in text:
                number = value
                break
    if number and 1 <= number <= 12:
        return f"Grade {number}", number + 2
    return clean_text(value), 99


def split_name(full_name):
    parts = clean_text(full_name).split()
    if not parts:
        return "", "", ""
    if len(parts) == 1:
        return parts[0], "", "Unknown"
    if len(parts) == 2:
        return parts[0], "", parts[1]
    return parts[0], " ".join(parts[1:-1]), parts[-1]


def locate_headers(sheet):
    for row_number in range(1, min(sheet.max_row, 40) + 1):
        mapping = {}
        for column_number, cell in enumerate(sheet[row_number], start=1):
            normalized = normalize_header(cell.value)
            for field, aliases in HEADER_ALIASES.items():
                if normalized in aliases:
                    mapping[field] = column_number
        if "student_name" in mapping and "grade" in mapping:
            return row_number, mapping
    raise ValueError("Could not find the sponsorship table headers in the workbook.")


def preview_workbook(file_path):
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    sheet = workbook["Master"] if "Master" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    header_row, columns = locate_headers(sheet)
    records = []
    for row_number in range(header_row + 1, sheet.max_row + 1):
        def value(field):
            column = columns.get(field)
            return sheet.cell(row_number, column).value if column else None

        student_name = clean_text(value("student_name"))
        if not student_name or student_name.lower() in {"student name", "total"}:
            continue
        dob = parse_excel_date(value("date_of_birth"))
        gender = normalize_gender(value("gender"))
        grade_name, grade_order = normalize_grade(value("grade"))
        sponsor_name = clean_text(value("sponsor_name"))
        errors = []
        warnings = []
        if not dob:
            errors.append("Missing or invalid date of birth")
        if not gender:
            errors.append("Missing or invalid gender")
        if not grade_name:
            errors.append("Missing grade")
        guardian_name = clean_text(value("guardian_name"))
        if not guardian_name:
            warnings.append("Guardian name is missing")
        if not sponsor_name:
            warnings.append("Student will be imported as unsponsored")
        records.append({
            "row_number": row_number,
            "student_name": student_name,
            "status": normalize_status(value("status")),
            "gender": gender,
            "date_of_birth": dob.isoformat() if dob else "",
            "guardian_name": guardian_name,
            "guardian_address": clean_text(value("guardian_address")),
            "sponsor_name": sponsor_name,
            "grade": grade_name,
            "grade_order": grade_order,
            "comments": clean_text(value("comments")),
            "errors": errors,
            "warnings": warnings,
            "valid": not errors,
        })
    return records


def track(batch, obj, created):
    if created:
        ImportedObject.objects.create(
            batch=batch, app_label=obj._meta.app_label, model_name=obj._meta.model_name,
            object_id=str(obj.pk), object_repr=str(obj)[:255]
        )


def import_records(batch, user):
    summary = {
        "students_created": 0, "students_existing": 0, "guardians_created": 0,
        "sponsors_created": 0, "grades_created": 0, "classes_created": 0,
        "enrollments_created": 0, "sponsorships_created": 0, "skipped": 0, "errors": [],
    }
    successful = 0
    failed = 0
    with transaction.atomic():
        for row in batch.preview_data:
            if not row.get("valid"):
                failed += 1
                summary["skipped"] += 1
                summary["errors"].append({"row": row["row_number"], "errors": row["errors"]})
                continue
            try:
                first, middle, last = split_name(row["student_name"])
                dob = date.fromisoformat(row["date_of_birth"])
                guardian = None
                if row["guardian_name"]:
                    guardian, created = Guardian.objects.get_or_create(
                        name__iexact=row["guardian_name"],
                        defaults={
                            "name": row["guardian_name"], "relationship": "Guardian",
                            "phone": "", "address": row["guardian_address"], "active": True,
                        },
                    )
                    track(batch, guardian, created)
                    summary["guardians_created"] += int(created)
                student = Student.objects.filter(
                    first_name__iexact=first, last_name__iexact=last, date_of_birth=dob
                ).first()
                student_created = False
                if not student:
                    student = Student.objects.create(
                        first_name=first, middle_name=middle, last_name=last, gender=row["gender"],
                        date_of_birth=dob, address=row["guardian_address"] or "Not provided",
                        guardian=guardian, admission_date=batch.academic_year.start_date,
                        previous_school="", is_active=True,
                    )
                    student_created = True
                    track(batch, student, True)
                    summary["students_created"] += 1
                else:
                    summary["students_existing"] += 1
                    if guardian and not student.guardian_id:
                        student.guardian = guardian
                        student.save(update_fields=["guardian"])
                grade, grade_created = GradeLevel.objects.get_or_create(
                    name__iexact=row["grade"],
                    defaults={"name": row["grade"], "order": row["grade_order"], "active": True},
                )
                track(batch, grade, grade_created)
                summary["grades_created"] += int(grade_created)
                class_section, class_created = ClassSection.objects.get_or_create(
                    grade=grade, name="A", defaults={"capacity": 60}
                )
                track(batch, class_section, class_created)
                summary["classes_created"] += int(class_created)
                enrollment_defaults = {"class_section": class_section, "active": True}
                if any(f.name == "admission_type" for f in Enrollment._meta.fields):
                    enrollment_defaults["admission_type"] = row["status"]
                enrollment, enrollment_created = Enrollment.objects.get_or_create(
                    student=student, academic_year=batch.academic_year, defaults=enrollment_defaults
                )
                track(batch, enrollment, enrollment_created)
                summary["enrollments_created"] += int(enrollment_created)
                sponsor = None
                if row["sponsor_name"]:
                    sponsor, sponsor_created = Sponsor.objects.get_or_create(
                        name__iexact=row["sponsor_name"],
                        defaults={"name": row["sponsor_name"], "active": True},
                    )
                    track(batch, sponsor, sponsor_created)
                    summary["sponsors_created"] += int(sponsor_created)
                sponsorship, sponsorship_created = StudentSponsorship.objects.get_or_create(
                    student=student, academic_year=batch.academic_year,
                    defaults={
                        "sponsor": sponsor,
                        "funding_status": "SPONSORED" if sponsor else "UNSPONSORED",
                        "coverage_type": "FULL", "coverage_value": Decimal("0.00"),
                        "notes": row["comments"], "active": True, "recorded_by": user,
                    },
                )
                track(batch, sponsorship, sponsorship_created)
                summary["sponsorships_created"] += int(sponsorship_created)
                successful += 1
            except Exception as exc:
                failed += 1
                summary["errors"].append({"row": row["row_number"], "errors": [str(exc)]})
    return summary, successful, failed


def rollback_batch(batch):
    deleted = 0
    errors = []
    for item in batch.created_objects.all().order_by("-id"):
        try:
            model = apps.get_model(item.app_label, item.model_name)
            obj = model.objects.filter(pk=item.object_id).first()
            if obj:
                obj.delete()
                deleted += 1
        except Exception as exc:
            errors.append(str(exc))
    batch.status = batch.Status.ROLLED_BACK
    batch.rolled_back_at = timezone.now()
    batch.summary = {**batch.summary, "rollback_deleted": deleted, "rollback_errors": errors}
    batch.save(update_fields=["status", "rolled_back_at", "summary"])
    return deleted, errors

# ==========================================================
# EMPLOYEE / STAFF IMPORT SUPPORT
# ==========================================================
from employees.models import Department as EmployeeDepartment, Employee, Position


def split_employee_description(value):
    text = clean_text(value)
    for separator in (" / ", " - ", " – ", " — ", ": "):
        if separator in text:
            return tuple(clean_text(part) for part in text.rsplit(separator, 1))
    return text, "Staff Member"


def split_employee_name(full_name):
    parts = clean_text(full_name).split()
    if len(parts) < 2:
        return "", "", ""
    if len(parts) == 2:
        return parts[0], "", parts[1]
    return parts[0], " ".join(parts[1:-1]), parts[-1]


def infer_employee_metadata(position_name):
    lower = clean_text(position_name).lower()
    is_teacher = "teacher" in lower or "instructor" in lower

    if is_teacher:
        department = "Teaching Staff"
    elif any(word in lower for word in ("principal", "registrar", "dean", "accountant", "admin", "director", "coordinator", "secretary", "it")):
        department = "Administration"
    elif any(word in lower for word in ("nurse", "health", "clinic")):
        department = "Health Services"
    elif any(word in lower for word in ("security", "guard")):
        department = "Security"
    else:
        department = "Support Services"

    employment_type = "Part Time" if re.search(r"\bPT\b|PART.?TIME", position_name, re.I) else "Full Time"
    cleaned_position = re.sub(r"\s*[-/]\s*(FT|PT)\s*$", "", position_name, flags=re.I).strip()
    return department, cleaned_position or "Staff Member", employment_type, is_teacher


def preview_employee_workbook(file_path):
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    records = []

    for row_number in range(1, sheet.max_row + 1):
        values = [clean_text(sheet.cell(row_number, column).value) for column in range(1, min(sheet.max_column, 8) + 1)]
        candidates = [value for value in values if value and len(value.split()) >= 2 and not value.isdigit()]
        if not candidates:
            continue

        raw_value = candidates[-1]
        if any(marker in raw_value.lower() for marker in ("employee name", "staff name", "name/position", "description")):
            continue

        full_name, position_name = split_employee_description(raw_value)
        first_name, middle_name, last_name = split_employee_name(full_name)
        if not first_name or not last_name:
            continue

        department_name, position_name, employment_type, is_teacher = infer_employee_metadata(position_name)
        duplicate = Employee.objects.filter(
            first_name__iexact=first_name,
            middle_name__iexact=middle_name,
            last_name__iexact=last_name,
        ).exists()

        records.append({
            "row_number": row_number,
            "full_name": full_name,
            "first_name": first_name,
            "middle_name": middle_name,
            "last_name": last_name,
            "position_name": position_name,
            "department_name": department_name,
            "employment_type": employment_type,
            "is_teacher": is_teacher,
            "duplicate": duplicate,
            "warnings": ["Complete gender, phone, address, date of birth and qualification after import."],
            "errors": [],
            "valid": True,
        })

    if not records:
        raise ValueError("No employee records were detected in the workbook.")
    return records


@transaction.atomic
def import_employee_records(batch, user):
    summary = {
        "employees_created": 0,
        "employees_existing": 0,
        "departments_created": 0,
        "positions_created": 0,
        "errors": [],
    }
    successful = 0
    failed = 0

    for row in batch.preview_data:
        try:
            existing = Employee.objects.filter(
                first_name__iexact=row["first_name"],
                middle_name__iexact=row["middle_name"],
                last_name__iexact=row["last_name"],
            ).first()
            if existing:
                summary["employees_existing"] += 1
                successful += 1
                continue

            department, created = EmployeeDepartment.objects.get_or_create(
                name__iexact=row["department_name"],
                defaults={"name": row["department_name"], "active": True},
            )
            track(batch, department, created)
            summary["departments_created"] += int(created)

            position, created = Position.objects.get_or_create(
                name__iexact=row["position_name"],
                defaults={"name": row["position_name"], "active": True},
            )
            track(batch, position, created)
            summary["positions_created"] += int(created)

            employee = Employee.objects.create(
                first_name=row["first_name"],
                middle_name=row["middle_name"],
                last_name=row["last_name"],
                gender=Employee.Gender.OTHER,
                phone="Not provided",
                address="",
                department=department,
                position=position,
                employment_type=row["employment_type"],
                is_teacher=row["is_teacher"],
                status=Employee.Status.ACTIVE,
                active=True,
                notes=f"Imported from {batch.original_filename}; source row {row['row_number']}.",
            )
            track(batch, employee, True)
            summary["employees_created"] += 1
            successful += 1
        except Exception as exc:
            failed += 1
            summary["errors"].append({"row": row["row_number"], "errors": [str(exc)]})

    return summary, successful, failed
